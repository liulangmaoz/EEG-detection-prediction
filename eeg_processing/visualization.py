import os
import numpy as np
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from scipy import signal

def generate_chart1_html(data, fs, save_dir):
    """
    生成全局偏移脑电图的HTML
    :param data: 脑电数据
    :param fs: 采样率
    :param save_dir: 保存目录
    """
    print("正在生成图表1的HTML内容（全局偏移脑电图）...")
    n_samples, n_channels = data.shape
    
    # 配置Matplotlib支持中文
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei',
        'SimHei',
        'Noto Sans CJK SC',
        'Segoe UI Symbol',
        'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['font.family'] = 'sans-serif'
    
    # Matplotlib绘图
    plt.figure(figsize=(18, 8))
    # 定义每个通道的垂直偏移量
    offset_step = 10000
    for i in range(n_channels):
        plt.plot(data[:, i] - offset_step * i, linewidth=0.5)
    
    plt.title("01全局偏移脑电图 (Global Offset View)", fontsize=14)
    plt.xlabel(f"Time (s) (Total: {n_samples / fs:.1f}s)", fontsize=12)
    
    # 隐藏Y轴的数字刻度和标题
    plt.gca().get_yaxis().set_ticks([])
    plt.ylabel("")
    
    # 在每个通道对应的Y轴位置添加电极名称
    from .core import CHANNEL_NAMES
    for i in range(n_channels):
        y_pos = -offset_step * i
        plt.text(-n_samples * 0.02, y_pos, CHANNEL_NAMES[i],
                 fontsize=10, ha='right', va='center')
    
    # 调整X轴范围，给电极名称留出空间
    plt.xlim(-n_samples * 0.02, n_samples)
    plt.tight_layout()
    
    # 图像转base64编码，嵌入HTML
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close()
    
    # 生成HTML模板
    html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>全局偏移脑电图</title>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; }}
        img {{ width: 100%; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>全局偏移脑电图 (Global Offset View)</h2>
        <div class="desc">完整数据：{n_samples} 点 ({n_samples / fs:.1f}秒)，12通道</div>
        <img src="data:image/png;base64,{img_base64}" alt="全局偏移脑电图">
    </div>
</body>
</html>
    """
    os.makedirs(save_dir, exist_ok=True)
    chart1_path = os.path.join(save_dir, "01全局偏移脑电图.html")
    with open(chart1_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"图表1已保存至: {chart1_path}")

def plot_psd_comparison_html(raw_data, lfp_data, fs, save_dir, channel_idx):
    """
    生成PSD频域对比图的HTML
    :param raw_data: 原始数据
    :param lfp_data: 滤波后数据
    :param fs: 采样率
    :param save_dir: 保存目录
    :param channel_idx: 通道索引
    """
    print("正在生成PSD频域对比图的HTML内容...")
    if raw_data.shape != lfp_data.shape:
        raise ValueError("原始数据与滤波后数据形状不一致！")
    if channel_idx < 0 or channel_idx >= raw_data.shape[1]:
        raise ValueError(f"通道索引 {channel_idx} 超出有效范围！")
    
    # 提取单通道数据
    from .core import CHANNEL_NAMES
    channel_name = CHANNEL_NAMES[channel_idx] if channel_idx < len(CHANNEL_NAMES) else f"通道{channel_idx}"
    raw_signal = raw_data[:, channel_idx]
    clean_signal = lfp_data[:, channel_idx]
    
    # PSD计算
    f_raw, Pxx_raw = signal.welch(
        raw_signal, fs=fs, nperseg=min(1024, len(raw_signal)//2),
        scaling='density', window='hann'
    )
    f_clean, Pxx_clean = signal.welch(
        clean_signal, fs=fs, nperseg=min(1024, len(clean_signal)//2),
        scaling='density', window='hann'
    )
    
    # 配置字体
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei',
        'SimHei',
        'Noto Sans CJK SC',
        'Segoe UI Symbol',
        'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['font.family'] = 'sans-serif'
    
    # 绘图
    plt.figure(figsize=(16, 10))
    # 上子图：原始信号PSD
    plt.subplot(2, 1, 1)
    plt.plot(f_raw, Pxx_raw, color='darkgray', linewidth=1.5, label=f'原始信号 - {channel_name}')
    plt.axvline(x=1, color='red', linestyle='--', alpha=0.7, label='1Hz 带通下限')
    plt.axvline(x=499, color='red', linestyle='--', alpha=0.7, label='499Hz 带通上限')
    plt.axvline(x=50, color='orange', linestyle='--', alpha=0.9, label='50Hz 工频干扰')
    plt.title(f'电极{channel_idx + 1}（{channel_name}）- 原始信号功率谱密度（PSD）', fontsize=12)
    plt.ylabel('功率谱密度 ($V^2/Hz$)', fontsize=10)
    plt.ylim(0, 150)
    plt.legend(loc='upper right', fontsize=9)
    plt.grid(alpha=0.3, linestyle='-', linewidth=0.5)
    plt.xlim(0, 100)
    
    # 下子图：滤波后信号PSD
    plt.subplot(2, 1, 2)
    plt.plot(f_clean, Pxx_clean, color='royalblue', linewidth=1.5, label=f'滤波后信号 - {channel_name}')
    plt.axvline(x=1, color='red', linestyle='--', alpha=0.7)
    plt.axvline(x=499, color='red', linestyle='--', alpha=0.7)
    plt.axvline(x=50, color='orange', linestyle='--', alpha=0.9)
    plt.title(f'电极{channel_idx + 1}（{channel_name}）- 滤波后信号功率谱密度（PSD）', fontsize=12)
    plt.xlabel('频率 (Hz)', fontsize=10)
    plt.ylabel('功率谱密度 ($V^2/Hz$)', fontsize=10)
    plt.ylim(0, 150)
    plt.legend(loc='upper right', fontsize=9)
    plt.grid(alpha=0.3, linestyle='-', linewidth=0.5)
    plt.xlim(0,100)
    
    plt.tight_layout()
    
    # 图像转base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close()
    
    # 生成HTML
    html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>02频域对比图</title>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; }}
        img {{ width: 100%; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>02频域对比图 - 电极{channel_idx + 1}（{channel_name}）</h2>
        <div class="desc">采样率：{fs}Hz | 带通范围：1Hz ~ 499Hz | 工频干扰：50Hz | Y轴范围：0~1000</div>
        <img src="data:image/png;base64,{img_base64}" alt="PSD频域对比图">
    </div>
</body>
</html>
    """
    
    # 保存HTML
    os.makedirs(save_dir, exist_ok=True)
    psd_chart_path = os.path.join(save_dir, "02频域对比图.html")
    with open(psd_chart_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"PSD频域对比图已保存至: {psd_chart_path}")

def generate_spike_detection_report(lfp_data, channel_stats_info, fs, save_dir, channel_idx, threshold_sd=3):
    """
    生成棘波检测报告
    :param lfp_data: 脑电数据
    :param channel_stats_info: 通道统计信息
    :param fs: 采样率
    :param save_dir: 保存目录
    :param channel_idx: 通道索引
    :param threshold_sd: 阈值倍数
    """
    print("正在生成棘波检测报告的HTML内容...")
    # 从通道统计字典中提取专属基线均值/标准差
    required_stats_keys = ['channel_baseline_means', 'channel_baseline_stds']
    for key in required_stats_keys:
        if key not in channel_stats_info:
            raise KeyError(f"通道统计字典缺少必要键：{key}，请确保字典包含基线统计数据")
    
    channel_baseline_means = channel_stats_info['channel_baseline_means']
    channel_baseline_stds = channel_stats_info['channel_baseline_stds']
    
    # 数据合法性校验
    if channel_idx < 0 or channel_idx >= lfp_data.shape[1]:
        raise ValueError(f"通道索引 {channel_idx} 超出有效范围（0 ~ {lfp_data.shape[1] - 1}）")
    if channel_idx < 0 or channel_idx >= len(channel_baseline_means) or channel_idx >= len(channel_baseline_stds):
        raise ValueError(f"通道索引 {channel_idx} 超出基线统计数据有效范围")
    if len(lfp_data.shape) != 2:
        raise ValueError("滤波后脑电数据应为二维数组，shape=(n_samples, n_channels)")
    
    fs = int(fs)
    lfp = lfp_data[:, channel_idx].copy()
    total_duration = len(lfp) / fs
    t = np.linspace(0, total_duration, len(lfp))
    
    # 提取通道专属基线
    baseline = channel_baseline_means[channel_idx]
    sd = channel_baseline_stds[channel_idx]
    
    # 跳过标准差为0的异常通道
    if sd == 0:
        from .core import CHANNEL_NAMES
        ch_name = CHANNEL_NAMES[channel_idx] if channel_idx < len(CHANNEL_NAMES) else f"通道{channel_idx}"
        print(f"警告：通道{ch_name}基线标准差为0，无法进行棘波检测，跳过该通道报告生成")
        return
    
    # 计算正负阈值
    positive_threshold = baseline + threshold_sd * sd
    negative_threshold = baseline - threshold_sd * sd
    
    # 打印阈值调试信息
    from .core import CHANNEL_NAMES
    ch_name = CHANNEL_NAMES[channel_idx] if channel_idx < len(CHANNEL_NAMES) else f"通道{channel_idx}"
    print(f"调试信息 - 通道{ch_name}：")
    print(f"  基线值(baseline)：{baseline:.2f}")
    print(f"  标准差(sd)：{sd:.2f}")
    print(f"  正向阈值(+{threshold_sd}σ)：{positive_threshold:.2f}")
    print(f"  反向阈值(-{threshold_sd}σ)：{negative_threshold:.2f}")
    
    distance = int(0.2 * fs)
    
    # 检测正向棘波
    peaks_h, props_h = signal.find_peaks(lfp, height=positive_threshold, distance=distance)
    widths_h = signal.peak_widths(lfp, peaks_h, rel_height=0.5)[0] / fs
    valid_h = widths_h < 0.1
    peaks_h = peaks_h[valid_h]
    pks_h = lfp[peaks_h] if len(peaks_h) > 0 else np.array([])
    
    # 检测反向棘波
    lfp_inv = -lfp
    peaks_l, props_l = signal.find_peaks(lfp_inv, height=positive_threshold, distance=distance)
    widths_l = signal.peak_widths(lfp_inv, peaks_l, rel_height=0.5)[0] / fs
    valid_l = widths_l < 0.1
    peaks_l = peaks_l[valid_l]
    pks_l = -lfp_inv[peaks_l] if len(peaks_l) > 0 else np.array([])
    
    # 打印检测结果
    print(f"提示：通道{ch_name}检测到 {len(pks_h)} 个正向棘波，{len(pks_l)} 个反向棘波")
    
    # 配置Matplotlib字体
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei',
        'SimHei',
        'Noto Sans CJK SC',
        'Segoe UI Symbol',
        'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['font.family'] = 'sans-serif'
    
    # 绘图
    fig, ax = plt.subplots(1, 1, figsize=(20, 6), facecolor='w')
    # 绘制脑电信号
    ax.plot(t, lfp, 'k-', linewidth=0.6, label='脑电信号')
    
    # 绘制正向棘波标记
    if len(peaks_h) > 0:
        ax.plot(t[peaks_h], pks_h, 'r*', markersize=8, label='正向棘波')
    
    # 绘制反向棘波标记
    if len(peaks_l) > 0:
        ax.plot(t[peaks_l], pks_l, 'b*', markersize=8, label='反向棘波')
    
    # 绘制正负阈值线
    ax.plot(t, np.ones_like(t) * positive_threshold, '--', color='r', linewidth=1.2,
            label=f'正向阈值 (+{threshold_sd}σ)')
    ax.plot(t, np.ones_like(t) * negative_threshold, '--', color='r', linewidth=1.2,
            label=f'反向阈值 (-{threshold_sd}σ)')
    
    # 设置图表标签和样式
    ax.set_title(f'发作间期棘波检测 (全时长 {total_duration:.1f}s) - 通道{ch_name}', fontsize=14)
    ax.set_xlabel('时间 (s)', fontsize=14)
    ax.set_ylabel('幅值', fontsize=14)
    ax.set_xlim(0, total_duration)
    ax.legend(loc='upper right', fontsize=10)
    ax.grid(alpha=0.3)
    
    # 自动调整Y轴范围
    ax.relim()
    ax.autoscale_view(True, True, True)
    
    plt.tight_layout()
    
    # 图像转base64编码
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)
    
    # 生成统一风格HTML模板
    desc = f"通道 {ch_name} | 全时长 {total_duration:.1f}s | 阈值 {threshold_sd}σ | 检测到正向棘波 {len(pks_h)} 个，反向棘波 {len(pks_l)} 个。"
    html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>03棘波检测</title>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; line-height: 1.6; }}
        img {{ width: 100%; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>03棘波检测 - 通道{ch_name}</h2>
        <div class="desc">{desc}</div>
        <img src="data:image/png;base64,{img_base64}" alt="发作间期棘波检测图">
    </div>
</body>
</html>
    """
    
    # 保存HTML文件
    os.makedirs(save_dir, exist_ok=True)
    spike_report_path = os.path.join(save_dir, "03棘波检测.html")
    with open(spike_report_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"棘波检测报告已保存至: {spike_report_path}")

def generate_seizure_detection_report(channel_stats_info, global_info, lfp_data, fs, save_dir):
    """
    生成发作检测报告
    :param channel_stats_info: 通道统计信息
    :param global_info: 全局信息
    :param lfp_data: 脑电数据
    :param fs: 采样率
    :param save_dir: 保存目录
    """
    print("正在生成发作检测分析报告...")
    
    # 导入CHANNEL_NAMES
    from .core import CHANNEL_NAMES
    
    # 从现有字典中提取所需数据
    required_global_keys = ['valid_global_events', 'global_energy']
    for key in required_global_keys:
        if key not in global_info:
            raise KeyError(f"全局信息字典缺少必要键：{key}，请确保字典包含该键")
    seizure_events = global_info['valid_global_events']
    global_energy = global_info['global_energy']
    
    # 提取发作检测阈值
    if 'global_threshold' in global_info:
        threshold = global_info['global_threshold']
    elif 'seizure_detection_threshold' in channel_stats_info:
        threshold = channel_stats_info['seizure_detection_threshold']
    else:
        raise KeyError("未找到发作检测阈值，请确保global_info或channel_stats_info中包含阈值数据")
    
    # 数据合法性校验
    if len(lfp_data.shape) != 2:
        raise ValueError("脑电数据应为二维数组，shape=(n_samples, n_channels) 或 (n_channels, n_samples)")
    # 适配两种数据形状
    if lfp_data.shape[0] == len(CHANNEL_NAMES) and lfp_data.shape[1] > lfp_data.shape[0]:
        pass
    elif lfp_data.shape[1] == len(CHANNEL_NAMES) and lfp_data.shape[0] > lfp_data.shape[1]:
        lfp_data = lfp_data.T
    else:
        raise ValueError("脑电数据形状与通道数量不匹配，请检查原始数据格式")
    
    # 校验全局能量长度与脑电数据采样点数量匹配
    n_samples = lfp_data.shape[1]
    if len(global_energy) != n_samples:
        raise ValueError(f"全局能量数据长度({len(global_energy)})与脑电数据采样点数量({n_samples})不匹配")
    
    # 校验保存目录
    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)
    
    # 配置Matplotlib字体
    plt.rcParams['font.sans-serif'] = [
        'Microsoft YaHei',
        'SimHei',
        'Noto Sans CJK SC',
        'Segoe UI Symbol',
        'DejaVu Sans'
    ]
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['font.family'] = 'sans-serif'
    
    # 生成时间轴
    time_axis = np.arange(n_samples) / fs
    
    # 绘制双图布局
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(20, 10), gridspec_kw={'height_ratios': [2, 1]}, sharex=True)
    
    # 上图：全局能量曲线 + 阈值 + 发作标注
    ax1.plot(time_axis, global_energy, label='全局平均能量', color='tab:blue', linewidth=1)
    ax1.axhline(threshold, color='r', linestyle='--', linewidth=2, label='发作阈值 (均值+3σ)')
    
    # 标注检测到的发作段
    for i, ev in enumerate(seizure_events):
        start_time = ev['start_time']
        end_time = ev['end_time']
        ax1.axvspan(start_time, end_time, color='orange', alpha=0.3, label='检测到的发作' if i == 0 else "")
        # 顶部标注发作编号
        ax1.text((start_time + end_time) / 2, np.max(global_energy), f"Seizure {i + 1}",
                 ha='center', va='bottom', fontsize=10, color='darkred')
    
    ax1.set_title("基于短时能量的癫痫大发作检测", fontsize=14)
    ax1.set_ylabel("能量 (RMS)", fontsize=12)
    ax1.legend(loc='upper right', fontsize=10)
    ax1.grid(alpha=0.3)
    
    # 下图：目标通道 (R_CA1) 参考信号
    from .core import CHANNEL_NAMES
    target_ch_idx = 2  # R_CA1
    target_sig = lfp_data[target_ch_idx, :]
    ax2.plot(time_axis, target_sig, 'k-', linewidth=0.6)
    ax2.set_title(f"参考信号: {CHANNEL_NAMES[target_ch_idx]}", fontsize=12)
    ax2.set_xlabel("时间 (s)", fontsize=12)
    ax2.set_ylabel("幅值", fontsize=12)
    ax2.grid(alpha=0.3)
    
    plt.tight_layout()
    
    # 图像转base64编码
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close(fig)
    
    # 生成统一风格HTML模板
    desc = f"检测阈值: {threshold:.2f} | 检测到 {len(seizure_events)} 次大发作事件"
    html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>04大发作检测</title>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; line-height: 1.6; font-size: 16px; }}
        img {{ width: 100%; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>04大发作检测</h2>
        <div class="desc">{desc}</div>
        <img src="data:image/png;base64,{img_base64}" alt="基于短时能量的癫痫大发作检测图">
    </div>
</body>
</html>
    """
    
    # 保存HTML文件
    seizure_report_path = os.path.join(save_dir, "04_大发作检测.html")
    with open(seizure_report_path, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"大发作检测报告已保存至: {seizure_report_path}")

def generate_multi_channel_timefreq_report(lfp_data, fs, save_dir, channel_names, font_cn, font_en):
    """
    生成多通道时频分析报告
    :param lfp_data: 脑电数据
    :param fs: 采样率
    :param save_dir: 保存目录
    :param channel_names: 通道名称
    :param font_cn: 中文字体
    :param font_en: 英文字体
    """
    n_samples, n_channels = lfp_data.shape
    # 调整子图布局：12个通道用6行2列
    fig, axes = plt.subplots(6, 2, figsize=(25, 40), facecolor='w')
    fig.suptitle('多通道时频分析 (Time-Frequency Analysis)', fontsize=20, fontweight='bold', fontproperties=font_cn)
    
    # 核心参数
    Nw = 1024
    noverlap = int(Nw * 0.8)
    nfft = 2048
    window = signal.windows.hamming(Nw)
    high_fre = 200
    vmin_percentile = 10
    vmax_percentile = 90
    
    # 计算时频图
    for ch_idx in range(n_channels):
        row = ch_idx // 2
        col = ch_idx % 2
        ax = axes[row, col]
        lfp = lfp_data[:, ch_idx].copy()
        total_duration = len(lfp) / fs
        
        # 计算时频图
        f, t, Sxx = signal.spectrogram(
            lfp,
            fs=fs,
            window=window,
            noverlap=noverlap,
            nfft=nfft,
            scaling='density',
            nperseg=Nw
        )
        
        # 过滤频率
        idx = np.where(f <= high_fre)[0][-1]
        f_filtered = f[:idx + 1]
        Sxx_filtered = Sxx[:idx + 1, :]
        Sxx_filtered = np.clip(Sxx_filtered, 1e-10, np.inf)
        Sxx_log = 10 * np.log10(Sxx_filtered)
        
        # 绘图
        im = ax.imshow(
            Sxx_log,
            extent=[t.min(), t.max(), f_filtered.min(), f_filtered.max()],
            aspect='auto',
            cmap='jet',
            vmin=np.percentile(Sxx_log, vmin_percentile),
            vmax=np.percentile(Sxx_log, vmax_percentile)
        )
        
        # 子图配置
        ax.set_title(
            f'癫痫发作时频分析 (全时长 {total_duration:.1f}s) - 通道{channel_names[ch_idx]}',
            fontsize=14,
            fontproperties=font_cn
        )
        ax.set_xlabel('时间 (s)', fontsize=12, fontproperties=font_cn)
        ax.set_ylabel('频率 (Hz)', fontsize=12, fontproperties=font_cn)
        ax.set_ylim(0, high_fre)
        ax.grid(True, linestyle='--', alpha=0.5)
        
        # 颜色条配置
        cbar = fig.colorbar(im, ax=ax, orientation='vertical', shrink=0.8)
        cbar.set_label('功率谱密度 (dB/Hz)', fontsize=12, fontproperties=font_cn)
        # 统一字体
        for label in ax.get_xticklabels() + ax.get_yticklabels() + cbar.ax.get_yticklabels():
            label.set_fontproperties(font_en)
    
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    
    # 保存为 Base64 并生成HTML
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
    buffer.seek(0)
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    plt.close()
    
    html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>多通道时频分析</title>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; }}
        img {{ width: 100%; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>多通道时频分析</h2>
        <div class="desc">展示所有12个电极通道在10分钟内的时频能量变化</div>
        <img src="data:image/png;base64,{img_base64}" alt="多通道时频分析">
    </div>
</body>
</html>
    """
    
    # 确保保存目录存在
    save_dir_path = os.path.dirname(save_dir)
    os.makedirs(save_dir_path, exist_ok=True)
    # 写入HTML文件
    with open(save_dir, "w", encoding="utf-8") as f:
        f.write(html_template)
    print(f"多通道时频分析报告已保存至: {save_dir}")

# 保存结果函数
def save_raw_data_view(raw_data, channel_names, save_dir):
    """
    保存原始数据查看
    :param raw_data: 原始数据
    :param channel_names: 通道名称
    :param save_dir: 保存目录
    """
    # 创建保存路径
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "01数据查看.xlsx")
    
    # 准备数据集基本信息
    n_samples, n_channels = raw_data.shape
    info_df = pd.DataFrame({
        "信息": [f"数据集形样本数: {n_samples}, 通道数: {n_channels}"]
    })
    
    # 准备通道统计信息
    channel_stats = []
    for i, ch_name in enumerate(channel_names):
        ch_data = raw_data[:, i]
        channel_stats.append({
            "电极通道": ch_name,
            "最小值": np.min(ch_data),
            "最大值": np.max(ch_data),
            "均值": np.mean(ch_data)
        })
    stats_df = pd.DataFrame(channel_stats)
    
    # 准备前20行数据
    n_rows = min(20, raw_data.shape[0])
    first_20_df = pd.DataFrame(
        raw_data[:n_rows, :],
        index=[f"第{row+1}行" for row in range(n_rows)],
        columns=channel_names
    )
    
    # 写入Excel
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # 写入基本信息
        info_df.to_excel(writer, sheet_name="数据查看", startrow=0, index=False, header=False)
        # 写入通道统计
        stats_df.to_excel(writer, sheet_name="数据查看", startrow=2, index=False)
        # 写入前20行数据
        first_20_df.to_excel(writer, sheet_name="数据查看", startrow=2 + len(stats_df) + 2)
        
        # 格式优化
        workbook = writer.book
        worksheet = writer.sheets["数据查看"]
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name="微软雅黑", size=10)
        
        # 遍历所有单元格设置格式
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        
        # 自动调整列宽
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 01数据查看.xlsx 已成功保存到: {excel_path}")

def save_seizure_detection(channel_stats_info, channel_spikes_list, earliest_channel_info, hfo_first_time, global_info, save_dir, fs, channel_names):
    """
    保存癫痫发作检测结果
    :param channel_stats_info: 通道统计信息
    :param channel_spikes_list: 通道棘波列表
    :param earliest_channel_info: 最早通道信息
    :param hfo_first_time: HFO首次出现时间
    :param global_info: 全局信息
    :param save_dir: 保存目录
    :param fs: 采样率
    :param channel_names: 通道名称
    """
    # 创建保存路径
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "02癫痫发作检测.xlsx")
    
    # 准备通道级数据
    channel_data = []
    for ch_idx, ch_name in enumerate(channel_names):
        # 提取基线数据并计算专属阈值
        ch_mean = channel_stats_info['channel_baseline_means'][ch_idx]
        ch_std = channel_stats_info['channel_baseline_stds'][ch_idx]
        ch_threshold = ch_mean + 3 * ch_std
        
        # 提取间期棘波数
        spike_count = channel_spikes_list[ch_idx] if ch_idx < len(channel_spikes_list) else 0
        
        # 从channel_valid_events统计该通道的发作次数
        ch_valid_events = channel_stats_info['channel_valid_events'][ch_idx]
        seizure_count = len(ch_valid_events)
        
        # 计算该通道的首次发作时间
        first_seizure_time = "无"
        if seizure_count > 0:
            earliest_event = min(ch_valid_events, key=lambda x: x['start_idx'])
            first_seizure_time = round(earliest_event['start_idx'] / fs, 3)
        
        # 提取首次HFO时间
        first_hfo_time = hfo_first_time.get(ch_name, "无")
        if isinstance(first_hfo_time, (np.float64, np.float32, float)):
            first_hfo_time = round(first_hfo_time, 3)
        
        # 组装单通道数据
        channel_data.append({
            "电极通道": ch_name,
            "基线均值": round(ch_mean, 3),
            "基线标准差": round(ch_std, 3),
            "专属阈值": round(ch_threshold, 3),
            "间期棘波数": spike_count,
            "发作次数": seizure_count,
            "首次发作时间（s）": first_seizure_time,
            "首次HFO时间（s）": first_hfo_time
        })
    df = pd.DataFrame(channel_data)
    
    # 准备全局参数行
    global_baseline_means = channel_stats_info['channel_baseline_means']
    global_baseline_stds = channel_stats_info['channel_baseline_stds']
    global_mean = round(np.mean(global_baseline_means), 3)
    global_std = round(np.mean(global_baseline_stds), 3)
    global_threshold = global_info.get('global_threshold', '无')
    if isinstance(global_threshold, (np.float64, np.float32, float)):
        global_threshold = round(global_threshold, 3)
    global_seizure_count = len(global_info.get('valid_global_events', []))
    earliest_channel = earliest_channel_info.get('earliest_channel', '无')
    
    global_params = {
        "电极通道": "全局参数",
        "基线均值": f"全局均值={global_mean}",
        "基线标准差": f"全局标准差={global_std}",
        "专属阈值": f"全局大发作阈值={global_threshold}",
        "间期棘波数": f"全局大发作次数={global_seizure_count}",
        "发作次数": f"最早发作通道={earliest_channel}",
        "首次发作时间（s）": "",
        "首次HFO时间（s）": ""
    }
    global_df = pd.DataFrame([global_params])
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="发作检测", index=False)
        global_df.to_excel(writer, sheet_name="发作检测", startrow=len(df) + 2, index=False)
        
        # 格式优化
        workbook = writer.book
        worksheet = writer.sheets["发作检测"]
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4) * 1.1
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 02癫痫发作检测.xlsx 已成功保存到: {excel_path}")

def save_band_energy_results(channel_band_results, save_dir):
    """
    保存频段能量结果
    :param channel_band_results: 通道频段能量结果
    :param save_dir: 保存目录
    """
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "03发作各阶段能量占比.xlsx")
    
    # 准备数据
    all_data = []
    for ch_name, ch_results in channel_band_results.items():
        for seizure_result in ch_results:
            seizure_id = seizure_result['seizure_id']
            for phase_result in seizure_result['phases']:
                band_info = phase_result['band_info']
                row = {
                    "电极通道": ch_name,
                    "发作ID": seizure_id,
                    "阶段": phase_result['phase_name'],
                    "DELTA能量": band_info['band_energy'].get('delta', 0),
                    "DELTA占比": band_info['band_relative_ratio'].get('delta', 0),
                    "THETA能量": band_info['band_energy'].get('theta', 0),
                    "THETA占比": band_info['band_relative_ratio'].get('theta', 0),
                    "ALPHA能量": band_info['band_energy'].get('alpha', 0),
                    "ALPHA占比": band_info['band_relative_ratio'].get('alpha', 0),
                    "BETA能量": band_info['band_energy'].get('beta', 0),
                    "BETA占比": band_info['band_relative_ratio'].get('beta', 0),
                    "GAMMA能量": band_info['band_energy'].get('gamma', 0),
                    "GAMMA占比": band_info['band_relative_ratio'].get('gamma', 0)
                }
                all_data.append(row)
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="能量占比", index=False)
        workbook = writer.book
        worksheet = writer.sheets["能量占比"]
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 03发作各阶段能量占比.xlsx 已保存到: {excel_path}")

def save_non_seizure_band_energy_results(channel_non_seizure_band_results, save_dir):
    """
    保存无发作频段能量结果
    :param channel_non_seizure_band_results: 无发作频段能量结果
    :param save_dir: 保存目录
    """
    # 创建保存目录
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "03无发作能量占比.xlsx")
    
    # 准备Excel数据
    all_data = []
    for ch_name, band_result in channel_non_seizure_band_results.items():
        band_info = band_result['band_info']
        row = {
            "电极通道": ch_name,
            "数据类型": "无癫痫发作",
            "提取起始时间(s)": band_result['start_sec'],
            "提取结束时间(s)": band_result['end_sec'],
            "提取时长(s)": band_result['duration_sec'],
            "DELTA能量": band_info['band_energy'].get('delta', 0),
            "DELTA占比": band_info['band_relative_ratio'].get('delta', 0),
            "THETA能量": band_info['band_energy'].get('theta', 0),
            "THETA占比": band_info['band_relative_ratio'].get('theta', 0),
            "ALPHA能量": band_info['band_energy'].get('alpha', 0),
            "ALPHA占比": band_info['band_relative_ratio'].get('alpha', 0),
            "BETA能量": band_info['band_energy'].get('beta', 0),
            "BETA占比": band_info['band_relative_ratio'].get('beta', 0),
            "GAMMA能量": band_info['band_energy'].get('gamma', 0),
            "GAMMA占比": band_info['band_relative_ratio'].get('gamma', 0)
        }
        all_data.append(row)
    
    # 转换为DataFrame
    if len(all_data) == 0:
        print("警告：无有效频段能量数据，无法生成Excel文件")
        return
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="无发作能量占比", index=False)
        workbook = writer.book
        worksheet = writer.sheets["无发作能量占比"]
        
        # 格式优化
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        
        # 自动调整列宽
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 03无发作能量占比.xlsx 已保存到: {excel_path}")

def save_phase_power_results(channel_phase_power_results, save_dir):
    """
    保存相位功率结果
    :param channel_phase_power_results: 通道相位功率结果
    :param save_dir: 保存目录
    """
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "04发作各阶段主频功率.xlsx")
    
    # 准备数据
    all_data = []
    for ch_name, ch_results in channel_phase_power_results.items():
        for seizure_result in ch_results:
            seizure_id = seizure_result['seizure_id']
            for phase_result in seizure_result['phases']:
                power_info = phase_result['power_info']
                row = {
                    "电极通道": ch_name,
                    "发作ID": seizure_id,
                    "阶段": phase_result['phase_name'],
                    "α主频": power_info['alpha']['freq'],
                    "α功率": power_info['alpha']['power'],
                    "α总功率": power_info['alpha']['total_power'],
                    "β主频": power_info['beta']['freq'],
                    "β功率": power_info['beta']['power'],
                    "β总功率": power_info['beta']['total_power'],
                    "α+β主频": power_info['combined']['freq'],
                    "α+β功率": power_info['combined']['power'],
                    "α+β总功率": power_info['combined']['total_power']
                }
                all_data.append(row)
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="主频功率", index=False)
        workbook = writer.book
        worksheet = writer.sheets["主频功率"]
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 04发作各阶段主频功率.xlsx 已保存到: {excel_path}")

def save_non_seizure_phase_power_results(channel_non_seizure_power_results, save_dir):
    """
    保存无发作相位功率结果
    :param channel_non_seizure_power_results: 无发作相位功率结果
    :param save_dir: 保存目录
    """
    # 创建保存目录
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "04无发作主频功率.xlsx")
    
    # 准备Excel数据
    all_data = []
    for ch_name, power_result in channel_non_seizure_power_results.items():
        power_info = power_result['power_info']
        row = {
            "电极通道": ch_name,
            "数据类型": "无癫痫发作",
            "提取起始时间(s)": power_result['start_sec'],
            "提取结束时间(s)": power_result['end_sec'],
            "提取时长(s)": power_result['duration_sec'],
            "α主频": power_info['alpha']['freq'],
            "α功率": power_info['alpha']['power'],
            "α总功率": power_info['alpha']['total_power'],
            "β主频": power_info['beta']['freq'],
            "β功率": power_info['beta']['power'],
            "β总功率": power_info['beta']['total_power'],
            "α+β主频": power_info['combined']['freq'],
            "α+β功率": power_info['combined']['power'],
            "α+β总功率": power_info['combined']['total_power']
        }
        all_data.append(row)
    
    # 转换为DataFrame
    if len(all_data) == 0:
        print("警告：无有效功率数据，无法生成Excel文件")
        return
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="无发作主频功率", index=False)
        workbook = writer.book
        worksheet = writer.sheets["无发作主频功率"]
        
        # 格式优化
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        
        # 自动调整列宽
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 04无发作主频功率.xlsx 已保存到: {excel_path}")

def save_spectral_entropy_results(channel_entropy_results, save_dir):
    """
    保存谱熵结果
    :param channel_entropy_results: 通道谱熵结果
    :param save_dir: 保存目录
    """
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "05发作各阶段功率谱熵.xlsx")
    
    # 准备数据
    all_data = []
    for ch_name, ch_results in channel_entropy_results.items():
        for seizure_result in ch_results:
            seizure_id = seizure_result['seizure_id']
            for phase_result in seizure_result['phases']:
                entropy = phase_result['entropy_info']
                row = {
                    "电极通道": ch_name,
                    "发作ID": seizure_id,
                    "阶段": phase_result['phase_name'],
                    "总功率谱熵": entropy.get('total_entropy', '无效'),
                    "归一化总熵": entropy.get('total_entropy_norm', '无效'),
                    "α波熵": entropy.get('alpha_entropy', '无效'),
                    "归一化α熵": entropy.get('alpha_entropy_norm', '无效'),
                    "β波熵": entropy.get('beta_entropy', '无效'),
                    "归一化β熵": entropy.get('beta_entropy_norm', '无效')
                }
                all_data.append(row)
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="功率谱熵", index=False)
        workbook = writer.book
        worksheet = writer.sheets["功率谱熵"]
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 05发作各阶段功率谱熵.xlsx 已保存到: {excel_path}")

def save_non_seizure_spectral_entropy_results(channel_non_seizure_entropy_results, save_dir):
    """
    保存无发作谱熵结果
    :param channel_non_seizure_entropy_results: 无发作谱熵结果
    :param save_dir: 保存目录
    """
    # 创建保存目录
    os.makedirs(save_dir, exist_ok=True)
    excel_path = os.path.join(save_dir, "05无发作功率谱熵.xlsx")
    
    # 准备Excel数据
    all_data = []
    for ch_name, entropy_result in channel_non_seizure_entropy_results.items():
        entropy_info = entropy_result['entropy_info']
        row = {
            "电极通道": ch_name,
            "数据类型": "无癫痫发作",
            "提取起始时间(s)": entropy_result['start_sec'],
            "提取结束时间(s)": entropy_result['end_sec'],
            "提取时长(s)": entropy_result['duration_sec'],
            "总功率谱熵": entropy_info.get('total_entropy', '无效'),
            "归一化总熵": entropy_info.get('total_entropy_norm', '无效'),
            "α波熵": entropy_info.get('alpha_entropy', '无效'),
            "归一化α熵": entropy_info.get('alpha_entropy_norm', '无效'),
            "β波熵": entropy_info.get('beta_entropy', '无效'),
            "归一化β熵": entropy_info.get('beta_entropy_norm', '无效')
        }
        all_data.append(row)
    
    # 转换为DataFrame
    if len(all_data) == 0:
        print("警告：无有效功率谱熵数据，无法生成Excel文件")
        return
    df = pd.DataFrame(all_data)
    
    # 写入Excel并优化格式
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="无发作功率谱熵", index=False)
        workbook = writer.book
        worksheet = writer.sheets["无发作功率谱熵"]
        
        # 格式优化
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name="微软雅黑", size=10)
        
        # 应用格式到所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = alignment
                cell.font = font
        
        # 自动调整列宽
        for col in worksheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = adjusted_width
    
    print(f"✅ 05无发作功率谱熵.xlsx 已保存到: {excel_path}")

# 交互式图表生成器
class InteractiveChartGenerator:
    @staticmethod
    def generate_echarts_sliding_window_html(plot_data, fs, directory, filename, channel_names):
        """
        生成交互式脑电图
        :param plot_data: 脑电数据
        :param fs: 采样率
        :param directory: 保存目录
        :param filename: 文件名
        :param channel_names: 通道名称
        """
        n_samples, n_channels = plot_data.shape
        total_duration = n_samples / fs
        
        # 确保保存目录存在
        os.makedirs(directory, exist_ok=True)
        
        # 生成完整的文件路径
        save_path = os.path.join(directory, filename)
        
        # 生成HTML
        html_template = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <title>05交互式脑电图</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{ margin: 0; padding: 20px; font-family: Arial, sans-serif; background-color: #f4f4f4; }}
        .chart-container {{ 
            background: white; padding: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); 
            border-radius: 5px; max-width: 1800px; margin: 0 auto;
        }}
        h2 {{ margin-top: 0; color: #333; text-align: center; }}
        .desc {{ text-align: center; color: #666; margin-bottom: 15px; }}
        #chart {{ width: 100%; height: 600px; }}
    </style>
</head>
<body>
    <div class="chart-container">
        <h2>05交互式脑电图</h2>
        <div class="desc">可拖动时间窗口查看不同时间段的脑电信号</div>
        <div id="chart"></div>
    </div>
    <script>
        var chart = echarts.init(document.getElementById('chart'));
        var data = {{
            channel_names: {channel_names},
            total_duration: {total_duration},
            fs: {fs}
        }};
        
        var option = {{
            title: {{
                text: '交互式脑电图',
                left: 'center'
            }},
            tooltip: {{
                trigger: 'axis',
                axisPointer: {{
                    type: 'cross'
                }}
            }},
            toolbox: {{
                feature: {{
                    dataZoom: {{
                        yAxisIndex: 'none'
                    }},
                    saveAsImage: {{}}
                }}
            }},
            dataZoom: [{{
                type: 'inside',
                start: 0,
                end: 100
            }}, {{
                start: 0,
                end: 100
            }}],
            legend: {{
                data: data.channel_names,
                orient: 'vertical',
                left: 'left'
            }},
            xAxis: {{
                type: 'value',
                name: '时间 (s)',
                min: 0,
                max: data.total_duration
            }},
            yAxis: {{
                type: 'value',
                name: '幅值'
            }},
            series: []
        }};
        
        chart.setOption(option);
        window.addEventListener('resize', function() {{
            chart.resize();
        }});
    </script>
</body>
</html>
        """
        
        # 写入HTML文件
        with open(save_path, "w", encoding="utf-8") as f:
            f.write(html_template)
        print(f"交互式脑电图已保存至: {save_path}")
